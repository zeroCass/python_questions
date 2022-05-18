import os 
from openpyxl import *
from openpyxl.styles import Font

dir = os.path.dirname(os.path.abspath(__file__))
root = os.path.abspath('.')

# formulas
wb = load_workbook(os.path.join(dir, 'examples.xlsx'), data_only=True)
sheet = wb.active
sheet['D1'] = '=SOMA(A1:C1)'
sheet['A1'] = 2
sheet['B1'] = 3
sheet['C1'] = 4
print(sheet['D1'].value)



# ajustando linhas e colunas
sheet.row_dimensions[1].height = 70
sheet.column_dimensions['C'].width = 20

# mesclando 
# rever é só usar o unmerge_cells()
sheet.merge_cells('C3:C5')


# congelar linhas
sheet_dir = os.path.join(root, 'automate_online-materials')
sheet1 = load_workbook(os.path.join(sheet_dir, 'produceSales.xlsx'))
ac_sheet = sheet1.active
# congela linhas que estao acima
ac_sheet.freeze_panes = 'A2'


# fonts
sheet['A1'].font = Font(size=24, italic=True, bold=True)

# salva a planilha
wb.save(os.path.join(dir,'examples.xlsx'))
sheet1.save(os.path.join(sheet_dir, 'produceSales_modified.xlsx'))


import openpyxl
import os

os.chdir(os.path.dirname(os.path.abspath(__file__)))

if os.path.exists('examples01.xlsx'):
    wb =openpyxl.load_workbook('examples01.xlsx')
else:
    wb = openpyxl.Workbook()
    wb.save('examples01.xlsx')

# plainha atual
sheet = wb.active
sheet.title = 'Sheet Random'
# cria uma planinha passando sua posicao e seu nome
wb.create_sheet(index=0, title='First Sheet')
wb.create_sheet(index=1, title='Second Sheet')

print(wb.sheetnames)
# uma forma de alterar o nome da planilha
wb['Sheet Random'].title = 'Third Sheet'
print(wb.sheetnames)
# remove a planilha informada
wb.remove(wb['Third Sheet'])
print(wb.sheetnames)

wb['Second Sheet']['A1'] = 'Hello World'
print(wb['Second Sheet']['A1'].value)


# salva /substitui uma pasta de trabalho 
wb.save('examples02.xlsx')
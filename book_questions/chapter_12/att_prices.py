'''
Projeto: Atualizando uma planilha

Nesse projeto, criaremos um programa para atualizar as células de uma
planilha de venda de produtos. Seu programa percorrerá a planilha, encontrará
tipos específicos de produtos e atualizará seus preços. Faça download dessa
planilha a partir de http://nostarch.com/automatestuff/. A figura 12.3 mostra a
aparência da planilha.

Cada linha representa uma venda individual. As colunas contêm o tipo de
produto vendido (A), o custo por libra (pound) desse produto (B), o número
de libras vendido (C) e a receita total dessa venda (D). A coluna TOTAL está
definida com a fórmula Excel =ROUND(B3*C3, 2), que multiplica o custo
por libra pelo número de libras vendido e arredonda o resultado para o
centavo mais próximo. Com essa fórmula, as células na coluna TOTAL serão
automaticamente atualizadas se houver uma mudança nas colunas B ou C.
Agora suponha que os preços do alho (garlic), do aipo (celery) e dos limões
(lemons) tenham sido inseridos incorretamente, deixando você com a tarefa
maçante de verificar milhares de linhas dessa planilha e atualizar o custo por
libra para todas as linhas contendo alho, aipo e limão. Não é possível
simplesmente pesquisar e substituir o preço, pois pode haver outros itens com
o mesmo preço que não devem ser erroneamente “corrigidos”. Se houver
milhares de linhas, essa tarefa demorará horas para ser feita manualmente. No
entanto, podemos criar um programa que faça isso em segundos.
'''

import os
import openpyxl

price_update = {
    'Celery': 1.19,
    'Garlic': 3.07,
    'Lemon': 1.27
}

sheet_dir = os.path.join(os.path.abspath('.'), 'automate_online-materials')
wb = openpyxl.load_workbook(os.path.join(sheet_dir, 'produceSales.xlsx'))
sheet = wb.active

for row in range(2, sheet.max_row + 1):
    # procura no dict se existe a key presente na coluna A na linha atual
    produce_name = sheet.cell(row=row, column=1).value
    if produce_name in price_update.keys():
        # substitue os valores
        sheet.cell(row=row, column=2).value = price_update[produce_name]
# salva um novo arquivo contendo os datos atualizados  
wb.save(os.path.join(sheet_dir, 'produceSales_update.xlsx'))

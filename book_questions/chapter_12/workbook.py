

import os
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.utils import column_index_from_string

raiz = os.path.join(os.path.abspath('.'), 'automate_online-materials') 

# cria obj workbook que representa a pasta de trabalho total, contendo todas as planinhas associadas
wb = openpyxl.load_workbook(os.path.join(raiz, 'example.xlsx'))
print(type(wb))
print(f'Plainhas: {wb.sheetnames}')

# pega determinada planilha
sheet = wb['Sheet3']
# pega a planilha atual
ac_sheet = wb.active
print(f'Planinha atual: {ac_sheet}')

# eh possivel obter os valores de uma determinada celula
print(ac_sheet['B1'].value)

# pode-se pega a linha e a coluna tbm
print('Linha ' + str(ac_sheet['C1'].row) + ', Coluna ' + str(ac_sheet['C1'].column) + ' is ' + str(ac_sheet['C1'].value)
     + '\nCoordenadas: ' + ac_sheet['C1'].coordinate +'\n')
    

# eh possivel acessar uma celular a partir de suas linhas e colunas informadas pela funcao cell(row=x, column=y)
for i in range(1, 8):
    print(i, ac_sheet.cell(row=i,column=2).value)
print(ac_sheet.max_column, ac_sheet.max_row)

# converte um numero para seu correspondente em uma letra nas colunas da planilha
print(get_column_letter(350))

# converte as letras das colunas para um valor inteiro
print(column_index_from_string('AA'))

# seleciona um intervalo de celulas
table = ac_sheet['A1:C7']
print(tuple(table))

# perce esse intervalo, seguindo a ordem da direia pra esquerda, de cima para baixo
for row in table:
    #print(f'Row: {row}')
    for cell_obj in row:
        print(cell_obj.coordinate, cell_obj.value, end=' ')
    print('\n--- End of row ---')
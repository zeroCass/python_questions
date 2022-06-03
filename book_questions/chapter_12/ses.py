import openpyxl
import os


os.chdir(os.path.dirname(os.path.abspath(__file__)))

wb = openpyxl.load_workbook('ADEQUACOES_MASCARAS_HORMONIOS_E_OUTROS.xlsx')
sheet_item_ex = wb['Planilha2']



list_item_ex = []
for row in range(2, sheet_item_ex.max_row + 1):

    obj = {}
    # CODIGO ANTIGO
    obj[sheet_item_ex.cell(row=1, column=1).value] = ''
    # CODIGO NOVO
    obj[sheet_item_ex.cell(row=1, column=2).value] = ''
    # ANALITO
    obj[sheet_item_ex.cell(row=1, column=3).value] = ''

    for col in range(1, sheet_item_ex.max_column + 1):
        obj[sheet_item_ex.cell(row=1, column=col).value] = sheet_item_ex.cell(row=row, column=col).value
    list_item_ex.append(obj)
    #print(obj)
#print(list_item_ex)



new_wb = openpyxl.Workbook()
new_sheet = new_wb.active
LOCAL_RECPT = []
wb2 = openpyxl.load_workbook('External_Codes_Imunologia_e_Hormonios.xlsx')
sheet_ext_codes = wb2['Imuno e Horm']
for row in range(2, sheet_ext_codes.max_row + 1):
    LOCAL_RECPT.append(sheet_ext_codes.cell(row=row, column=1).value)

LOCAL_RECPT = set(LOCAL_RECPT)



TO_LIST = []
_row = 1
with open('text.txt', 'w') as f:
    for local in LOCAL_RECPT:
        if local is None:
            continue

        # obj = {}
        # obj[sheet_ext_codes.cell(row=1, column=1).value] = ''
        # obj[sheet_ext_codes.cell(row=1, column=2).value] = ''
        # obj[sheet_ext_codes.cell(row=1, column=3).value] = ''
        # obj[sheet_ext_codes.cell(row=1, column=4).value] = ''
        # obj[sheet_ext_codes.cell(row=1, column=5).value] = ''
        # obj[''] = ''
        # obj[sheet_ext_codes.cell(row=1, column=7).value] = ''
        # obj[sheet_ext_codes.cell(row=1, column=8).value] = ''

        for exame in list_item_ex:
            obj = {}
            obj[sheet_ext_codes.cell(row=1, column=1).value] = local
            obj[sheet_ext_codes.cell(row=1, column=2).value] = ''
            obj[sheet_ext_codes.cell(row=1, column=3).value] = exame['CÓDIGO NOVO']
            obj[sheet_ext_codes.cell(row=1, column=4).value] = exame['ANALITOS']
            obj[sheet_ext_codes.cell(row=1, column=5).value] = '06/06/22'
            obj[''] = ''
            obj[sheet_ext_codes.cell(row=1, column=7).value] = exame['CÓDIGO ANTIGO']
            obj[sheet_ext_codes.cell(row=1, column=8).value] = '05/06/22'
            f.write(str(obj))
            TO_LIST.append(obj)

            new_sheet.cell(row=_row, column=1).value = local
            new_sheet.cell(row=_row, column=2).value = ''
            new_sheet.cell(row=_row, column=3).value = exame['CÓDIGO NOVO']
            new_sheet.cell(row=_row, column=4).value = exame['ANALITOS']
            new_sheet.cell(row=_row, column=5).value = '06/06/22'
            new_sheet.cell(row=_row, column=6).value = ''
            new_sheet.cell(row=_row, column=7).value = exame['CÓDIGO ANTIGO']
            new_sheet.cell(row=_row, column=8).value = '05/06/22'
            _row += 1
        _row += 2
        f.write('\n\n\n')
        

print(TO_LIST)
print(f'ROW: {_row}\n\n')
new_wb.save('new_wb.xlsx')





        


